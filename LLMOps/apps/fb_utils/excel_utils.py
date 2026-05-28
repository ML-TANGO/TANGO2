from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import Dict, List, Any
import datetime

def create_uptime_excel(uptime_data: Dict[str, Any], start_datetime: str, end_datetime: str) -> BytesIO:
    """
    업타임 데이터를 Excel 파일로 변환
    
    Args:
        uptime_data: 업타임 데이터 (workspaceUptimes, projectUptimes 포함)
        start_datetime: 조회 시작 시간
        end_datetime: 조회 종료 시간
    
    Returns:
        BytesIO: Excel 파일 바이너리 데이터
    """
    wb = Workbook()
    
    # 기본 시트 제거
    wb.remove(wb.active)
    
    # 스타일 정의
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 1. Workspace Uptimes 시트
    ws1 = wb.create_sheet("Workspace Uptimes")
    
    # 헤더 정보
    ws1['A1'] = f"Workspace Uptime Report"
    ws1['A2'] = f"Period: {start_datetime} ~ {end_datetime}"
    ws1['A3'] = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # 헤더 스타일 적용
    for row in range(1, 4):
        ws1[f'A{row}'].font = Font(bold=True)
    
    # 테이블 헤더 (5행부터 시작)
    headers = ["Workspace Name", "Global Uptime (seconds)", "Office Hour Uptime (seconds)", 
               "Global Uptime (hours)", "Office Hour Uptime (hours)"]
    
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 데이터 입력
    workspace_uptimes = uptime_data.get("workspaceUptimes", [])
    for row, workspace in enumerate(workspace_uptimes, 6):
        ws1.cell(row=row, column=1, value=workspace.get("workspaceName", ""))
        
        global_seconds = int(workspace.get("globalUptimeSeconds", 0))
        office_seconds = int(workspace.get("officeHourUptimeSeconds", 0))
        
        ws1.cell(row=row, column=2, value=global_seconds)
        ws1.cell(row=row, column=3, value=office_seconds)
        ws1.cell(row=row, column=4, value=round(global_seconds / 3600, 2))  # 시간 단위
        ws1.cell(row=row, column=5, value=round(office_seconds / 3600, 2))  # 시간 단위
    
    # 열 너비 자동 조정
    for col in range(1, 6):
        column_letter = get_column_letter(col)
        ws1.column_dimensions[column_letter].width = 20
    
    # 2. Project Uptimes 시트
    ws2 = wb.create_sheet("Project Uptimes")
    
    # 헤더 정보
    ws2['A1'] = f"Project Uptime Report"
    ws2['A2'] = f"Period: {start_datetime} ~ {end_datetime}"
    ws2['A3'] = f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    
    # 헤더 스타일 적용
    for row in range(1, 4):
        ws2[f'A{row}'].font = Font(bold=True)
    
    # 테이블 헤더 (5행부터 시작)
    project_headers = ["Workspace Name", "Type", "Project Name", "Global Uptime (seconds)", 
                      "Office Hour Uptime (seconds)", "Global Uptime (hours)", "Office Hour Uptime (hours)"]
    
    for col, header in enumerate(project_headers, 1):
        cell = ws2.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 데이터 입력
    project_uptimes = uptime_data.get("projectUptimes", [])
    for row, project in enumerate(project_uptimes, 6):
        ws2.cell(row=row, column=1, value=project.get("workspaceName", ""))
        ws2.cell(row=row, column=2, value=project.get("type", ""))
        ws2.cell(row=row, column=3, value=project.get("projectName", ""))
        
        global_seconds = int(project.get("globalUptimeSeconds", 0))
        office_seconds = int(project.get("officeHourUptimeSeconds", 0))
        
        ws2.cell(row=row, column=4, value=global_seconds)
        ws2.cell(row=row, column=5, value=office_seconds)
        ws2.cell(row=row, column=6, value=round(global_seconds / 3600, 2))  # 시간 단위
        ws2.cell(row=row, column=7, value=round(office_seconds / 3600, 2))  # 시간 단위
    
    # 열 너비 자동 조정
    for col in range(1, 8):
        column_letter = get_column_letter(col)
        ws2.column_dimensions[column_letter].width = 18
    
    # Excel 파일을 메모리에 저장
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

def format_filename(start_datetime: str, end_datetime: str) -> str:
    """
    파일명 생성
    
    Args:
        start_datetime: 시작 시간
        end_datetime: 종료 시간
    
    Returns:
        str: 파일명
    """
    # 날짜 형식 변환 (공백과 특수문자 제거)
    start_clean = start_datetime.replace(" ", "_").replace(":", "_")
    end_clean = end_datetime.replace(" ", "_").replace(":", "_")
    
    return f"uptime_{start_clean}_{end_clean}.xlsx" 